import openpyxl 
import os
import datetime
from fuzzywuzzy import fuzz

start_time = datetime.datetime.now()

os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\3rd Party\Master Files\Foreign Sources Reports")
wb = openpyxl.load_workbook("Foreign Sources Report 1208.xlsx")
foreignNames = openpyxl.load_workbook("_Foreign Sources Names.xlsx")
data = wb['Data']
foreignNameList = foreignNames.active

resultswb = openpyxl.Workbook()
resultsSheet = resultswb.active

count = 0
for i, ele in enumerate(list(foreignNameList.columns)[0]):
    #print(ele.value)
    for j, elej in enumerate(list(data.columns)[7]):
        #print(fuzz.token_set_ratio(ele.value, elej.value))
        if fuzz.token_set_ratio(ele.value, elej.value) > 90:
            count += 1
            for index, element in enumerate(list(data.rows)[j]):
                resultsSheet.cell(row=count, column=index+1).value = element.value
                print(element.value, end=" ")
            print('\n')
            
print(count, 'Elapsed: ',datetime.datetime.now() - start_time)
resultswb.save('Test Results_' + str(datetime.date.today()) + '.xlsx')

##########>>>>>>>>
""" 1) need to create list of names that have already been verified as noninternational
and compare that list to current list element to avoid reviewing duplicate noninternation orgs
    2) need to incorperate the sponsor data in a more clean/smooth way""" 